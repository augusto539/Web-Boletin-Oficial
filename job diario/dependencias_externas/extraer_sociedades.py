"""
Script para extraer información de sociedades del Boletín Oficial de Mendoza.
Usa pdfplumber para extraer texto y la API de Claude (Haiku) para parsear
los campos estructurados de forma robusta.

Modo de uso:
  - TEST_MODE = True   → procesa solo el primer PDF (verificar resultados)
  - TEST_MODE = False  → procesa todos los PDFs de la carpeta boletines/
"""

import json
import logging
import os
import re
import time
from pathlib import Path
from datetime import datetime

import anthropic
import pdfplumber
import pandas as pd

# ── Configuración ────────────────────────────────────────────────────────────
# Carpeta de PDFs y Excel de salida. Se pueden sobreescribir por variable de entorno
# (útil para corridas acotadas como un backfill parcial, sin editar el código):
#   BOLETIN_INPUT_DIR   -> carpeta de PDFs a procesar
#   BOLETIN_OUTPUT_FILE -> ruta del Excel de salida
#   BOLETIN_CHECKPOINT  -> ruta del checkpoint (conviene uno propio por corrida aislada)
BOLETINES_DIR       = Path(os.getenv("BOLETIN_INPUT_DIR", str(Path(__file__).parent / "PDFs" / "boletines")))
BOLETINES_PRUEBA    = Path(__file__).parent / "PDFs" / "prueba campos nuevos"
OUTPUT_FILE         = Path(os.getenv("BOLETIN_OUTPUT_FILE", str(Path(__file__).parent / "Resultados" / "Sociedades Mendoza 2023 - 2026.xlsx")))
LOGS_DIR            = Path(__file__).parent / "Logs"
CHECKPOINT_FILE     = Path(os.getenv("BOLETIN_CHECKPOINT", str(Path(__file__).parent / "checkpoint.json")))

LOGS_DIR.mkdir(exist_ok=True)
_LOG_FILE = LOGS_DIR / f"{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"

_FORMATO_LOG = "%(asctime)s [%(levelname)s] %(message)s"


class _FormatterColorConsola(logging.Formatter):
    """Colorea SOLO la salida en consola (el archivo de log queda en texto plano,
    para poder leerlo después sin códigos ANSI de por medio): la línea entera en
    rojo si es un error, amarillo si es un warning, verde si es un guardado de
    checkpoint."""
    _ROJO      = "\033[31m"
    _AMARILLO  = "\033[33m"
    _VERDE     = "\033[32m"
    _RESET     = "\033[0m"

    def format(self, record: logging.LogRecord) -> str:
        linea = super().format(record)
        if record.levelno >= logging.ERROR:
            return f"{self._ROJO}{linea}{self._RESET}"
        if record.levelno == logging.WARNING:
            return f"{self._AMARILLO}{linea}{self._RESET}"
        if "checkpoint guardado" in record.getMessage():
            return f"{self._VERDE}{linea}{self._RESET}"
        return linea


_handler_archivo = logging.FileHandler(_LOG_FILE, encoding="utf-8")
_handler_archivo.setFormatter(logging.Formatter(_FORMATO_LOG))

_handler_consola = logging.StreamHandler()
_handler_consola.setFormatter(_FormatterColorConsola(_FORMATO_LOG))

logging.basicConfig(
    level=logging.INFO,
    handlers=[_handler_archivo, _handler_consola],
)

TEST_MODE      = False             # True = carpeta de prueba; False = usa BOLETINES_DIR (o BOLETIN_INPUT_DIR)
# En modo prueba: salida y checkpoint separados para NO pisar el dataset real
if TEST_MODE:
    OUTPUT_FILE     = Path(__file__).parent / "Resultados" / "PRUEBA campos nuevos.xlsx"
    CHECKPOINT_FILE = Path(__file__).parent / "checkpoint_prueba.json"
MODEL          = "claude-haiku-4-5"   # Barato y suficiente para extracción estructurada
MAX_TOKENS     = 3072             # Subido a 3072: aún había truncados en bloques con muchos socios.
                                   # Con BLOQUES_POR_LLAMADA=2 el peor caso medido (~615 tok/bloque)
                                   # sigue con margen amplio (~1230 de 3072). Si se sube el agrupamiento,
                                   # revisar este margen.
MAX_REINTENTOS = 3                # Intentos ante errores de API (rate limit, timeout, conexión)
ESPERA_BASE    = 5                # Segundos base para backoff exponencial en errores de conexión
MAX_WORKERS    = 32                # PDFs procesados en paralelo (1 = secuencial)
BLOQUES_POR_LLAMADA = 2            # Bloques del boletín agrupados en una sola llamada a Claude.
                                   # Reduce el número de llamadas (y cuántas veces se paga la lectura
                                   # del caché del system prompt) sin cambiar qué se extrae de cada uno.
BLOQUES_EN_PARALELO = 6            # Copia usada por el job diario (1 PDF/día típico): la paralelización
                                   # de MAX_WORKERS es POR ARCHIVO y no ayuda con un solo PDF. Acá se
                                   # paralelizan además los grupos DENTRO de un mismo PDF (independientes
                                   # entre sí), ya que extraer_datos_claude()/get_client() son thread-safe
                                   # (_stats_lock, singleton del cliente) — ver procesar_pdf().

COLUMNAS = [
    "Nombre de la sociedad",
    "CUIT de la sociedad",
    "Tipo de sociedad",

    "Fecha del acto",
    "ID del boletín",
    "Fecha publicación en boletín",
    "Capital inicial",
    "Tipo de acto",
    "Descripción del acto",
    "Nombres de los socios",
    "Cargos de los socios",
    "Fechas de nacimiento de los socios",
    "Profesiones de los socios",
    "Objeto social",
    "DNI de los socios",
    "CUIT/CUIL de los socios",
    "Domicilios particulares de los socios",
    "Porcentaje de los socios",
    "Nombres de los socios jurídicos",
    "CUIT de los socios jurídicos",
    "Porcentaje de los socios jurídicos",
    "Nombres de los apoderados",
    "DNI de los apoderados",
    "Domicilio de la sociedad",
    "Calle y número de la sociedad",
    "Localidad de la sociedad",
    "Escribano interviniente",
    "Registro notarial",
    "Domicilio electrónico de la sociedad",
    "Domicilio electrónico de los socios",
]

# Títulos de secciones del boletín que vienen DESPUÉS de CONTRATOS SOCIALES.
# Se usan para detectar dónde termina la sección de sociedades (compartido por el
# path en tiempo real y el path de Batch API).
SECCIONES_SIGUIENTES = [
    "CONVOCATORIAS",
    "CONCURSOS Y QUIEBRAS",
    "IRRIGACION",
    "IRRIGACIÓN",
    "NOTIFICACIONES",
    "EDICTOS",
    "RESOLUCIONES",
    "DISPOSICIONES",
    "DECRETOS",
    "LICITACIONES",
    "CONCURSOS DE PRECIOS",
    "CONCURSO DE PRECIOS",
    "AVISOS OFICIALES",
    "REMATES JUDICIALES",
    "REMATES OFICIALES",
    "DIRECCION PROVINCIAL",
    "DIRECCIÓN PROVINCIAL",
    "ADMINISTRADORA PROVINCIAL",
]

# ── System prompt (se envía con cache_control para no cobrarlo N veces) ───────
SYSTEM_PROMPT = """Eres un asistente especializado en extraer datos estructurados de textos legales del Boletín Oficial de Mendoza, Argentina.

Recibirás el texto de uno o más bloques, cada uno correspondiente a una sociedad o entidad independiente (puede ser S.A., S.R.L., S.A.S., Asociación Civil, Sucursal Extranjera, Cooperativa, Fundación, Fideicomiso, Unión Transitoria u otro tipo). Si hay más de un bloque, vienen precedidos por un separador "=== BLOQUE N ===".

Tu tarea es extraer los siguientes campos de CADA bloque y devolver, por cada uno, ÚNICAMENTE un objeto JSON válido, sin texto adicional, sin bloques de código markdown. Si recibiste varios bloques, devolvé varios objetos JSON, uno a continuación del otro en el mismo orden en que aparecen los bloques (sin arreglo `[...]`, sin numerarlos ni comentar nada entre uno y otro) — cada bloque es independiente, no mezcles datos de un bloque con otro.

IMPORTANTE para ahorrar espacio: cada objeto JSON va en una sola línea, sin saltos de línea ni indentación entre claves, y OMITIENDO por completo las claves cuyo valor quedaría vacío (no las incluyas con "" — directamente no las escribas). La única excepción son "email_sociedad" y "email_socios", que SIEMPRE van presentes: con el email si figura, o con el string exacto "N/A" si no figura.

Campos a extraer de cada bloque:

{
  "nombre": "Nombre de la sociedad o entidad, sin prefijos de acto legal (CONSTITUCION DE, INSCRIPCION DE, etc.) y sin comillas. Solo el nombre. En formato Title Case.",
  "cuit_sociedad": "CUIT de la sociedad/entidad cuyo acto se publica, sin guiones (ej: 30712345678). En las CONSTITUCIONES normalmente NO figura (la sociedad aún no tiene CUIT): dejar vacío. En aumentos, modificaciones, cesiones, disoluciones, designación de autoridades, etc. casi siempre figura el CUIT de la sociedad en el encabezado: extraelo. NO confundir con las CUIT de los socios (cuit_socios) ni de los socios jurídicos.",
  "tipo": "Tipo societario: S.A. / S.R.L. / S.A.S. / Asociacion Civil / Sociedad Civil / Sucursal Extranjera / Cooperativa / Fundacion / Fideicomiso / Union Transitoria / Otro",

  "fecha_acto": "Fecha del acto que se publica, en formato DD/MM/YYYY. Para una constitución es la fecha del acto constitutivo; para un aumento o modificación es la fecha de la asamblea o reunión que lo aprobó; para una cesión es la fecha de la cesión; etc. Si figuran varias fechas, usá la del acto principal que se publica. Si no figura, dejar vacío.",
  "capital": "Capital social con símbolo $. Ejemplo: $500.000 o $1.200.000,00. Si no figura, dejar vacío.",
  "tipo_acto": "Tipo de publicación: Constitución / Aumento de capital / Modificación de estatuto / Disolución / Liquidación / Inscripción de sucursal / Transformación / Fusión / Escisión / Otro",
  "descripcion_acto": "Resumen en 1-2 frases, en lenguaje natural, de QUÉ es o qué resuelve este acto, con los datos concretos. OBLIGATORIO completarlo para TODO acto que NO sea una constitución — incluí convocatorias a asamblea, designación o renovación de autoridades, aumentos, cesiones, disoluciones, transformaciones, etc. Ejemplos: 'Convocatoria a Asamblea General Ordinaria para elección de comisión directiva y aprobación de balance.' / 'Se aumentó el capital social de $100.000 a $600.000.' / 'Juan Pérez cedió 20 cuotas a María Gómez.' / 'Se designó nuevo directorio: Presidente Ana Ruiz.' / 'Se modificó el objeto social para incluir actividad inmobiliaria.' SOLO en constituciones podés dejarlo vacío (los datos ya están en los campos estructurados). Es texto libre; priorizá nombres, montos, fechas y porcentajes concretos.",
  "nombres_socios": "Nombres completos de TODAS las personas físicas mencionadas: socios/accionistas Y también directores, administradores, síndicos, gerentes o miembros de comisión directiva que NO sean socios. Separados por punto y coma. NO incluyas acá a las sociedades/empresas que sean socias (esas van en socios_juridicos). Para Fideicomisos: incluir fiduciante(s) y/o beneficiario(s) que sean personas físicas. Para Uniones Transitorias: dejá esto vacío y listá las empresas integrantes en socios_juridicos.",
  "cargos_socios": "Cargo o rol de cada persona en el mismo orden que nombres_socios, separados por punto y coma. Valores posibles — S.A.S.: Socio, Administrador Titular, Administrador Suplente. S.A.: Socio, Presidente, Vicepresidente, Director Titular, Director Suplente, Síndico Titular, Síndico Suplente. S.R.L.: Socio, Gerente Titular, Gerente Suplente. Asociación Civil: Socio Fundador, Presidente, Secretario, Tesorero, Vocal Titular, Vocal Suplente, Revisor de Cuentas. Fideicomiso: Fiduciante, Beneficiario, Fiduciario. Si una persona tiene varios roles (ej: es socio Y administrador titular) unilos con ' / ' (ej: 'Socio / Administrador Titular'). Si el rol no queda claro: 'Socio'.",
  "fechas_nacimiento_socios": "Fechas de nacimiento de los socios separadas por punto y coma, en el mismo orden que nombres_socios, formato DD/MM/YYYY. Buscá patrones como 'nacido el DD/MM/YYYY', 'nacida el DD de mes de YYYY', 'de NN años de edad nacido el...', 'fecha de nacimiento: DD/MM/YYYY'. Si un socio no tiene fecha de nacimiento en el texto, dejá su posición vacía (ej: si hay 3 socios y solo el 2do tiene fecha: '; 15/03/1985; '). Si ninguno tiene: dejar vacío.",
  "profesiones_socios": "Profesiones u ocupaciones de los socios separadas por punto y coma. En minúscula. Para Fideicomisos y Uniones Transitorias: dejar vacío.",
  "objeto_social": "Descripción del objeto social de la empresa. LÍMITE ESTRICTO de 400 caracteres — contalos y cortá la frase antes de superarlo, no lo excedas.",
  "cuit_socios": "CUIT o CUIL de los socios PERSONAS FÍSICAS separados por punto y coma, sin guiones (ej: 20123456789; 27345678901). Las CUIT de sociedades socias NO van acá, van en cuit_socios_juridicos. Si no figuran: dejar vacío.",
  "dni_socios": "Documentos de identidad de los socios separados por punto y coma, sin puntos ni comas dentro del número. Para DNI argentino: solo el número (ej: 12345678). Para extranjeros: prefijo del tipo de documento seguido del número (ej: CI: 17623124-6 o PAS: ABC123456). Si no figuran documentos: dejar vacío.",
  "domicilios_socios": "Domicilios particulares o reales de los socios separados por punto y coma, en el mismo orden que nombres_socios. Buscá 'con domicilio en', 'domicilio real en', 'domiciliado en', 'domiciliada en'. Solo el domicilio personal de cada socio, NO el domicilio social de la empresa. Si un socio no tiene domicilio particular registrado, dejá su posición vacía. Si ninguno tiene: dejar vacío.",
  "porcentajes_socios": "Porcentaje de PARTICIPACIÓN (acciones/cuotas suscriptas) de cada socio PERSONA FÍSICA en el capital, en el mismo orden que nombres_socios, separados por punto y coma, con el signo % (ej: '50%; 50%' o '33,33%; 66,67%'). IMPORTANTE: si el texto da la cantidad de acciones/cuotas que suscribe cada socio (ej: 'suscribe 250 de las 500 acciones', o 'suscribe 1.000.000 acciones' siendo el capital de 1.000.000 acciones), CALCULÁ SIEMPRE el porcentaje (acciones del socio ÷ total, redondeá a 2 decimales). No te quedes con el campo vacío si hay datos para calcularlo. NO confundas participación/suscripción con INTEGRACIÓN (lo que cada socio paga, ej: 'integra el 25%'): la integración NO es el porcentaje de participación. Si el texto solo da el capital total sin desglose por socio (ni acciones ni % por cada uno), dejá vacío — NO inventes un reparto igualitario. Recordá: cada posición corresponde a un nombre de nombres_socios; los administradores que no son socios llevan posición vacía.",
  "socios_juridicos": "Nombres de las personas JURÍDICAS (otras sociedades/empresas) que son socias o accionistas de esta entidad, separados por punto y coma, en Title Case. Incluí sociedades como S.A., S.R.L., S.A.S., cooperativas, fundaciones u otras empresas que figuren como socias. NO incluyas acá a las personas físicas (esas van en nombres_socios). Para Uniones Transitorias: listá acá TODAS las empresas integrantes. Si no hay socios persona jurídica: dejar vacío.",
  "cuit_socios_juridicos": "CUIT de las sociedades socias listadas en socios_juridicos, en el mismo orden, separados por punto y coma, sin guiones (las CUIT de personas jurídicas suelen empezar en 30, 33 o 34). Si una sociedad socia no tiene CUIT en el texto, dejá su posición vacía. Si no hay socios persona jurídica: dejar vacío.",
  "porcentajes_socios_juridicos": "Porcentaje de participación en el capital de cada sociedad socia, en el mismo orden que socios_juridicos, separados por punto y coma, con el signo % (ej: '60%; 40%'). Mismas reglas que porcentajes_socios: si no está explícito pero hay acciones/cuotas y total, calculalo; si no se puede determinar, dejá la posición vacía. Si no hay socios persona jurídica: dejar vacío.",
  "apoderados": "Nombres completos de los apoderados, representantes o mandatarios de la sociedad (personas a quienes se les otorga PODER para actuar), separados por punto y coma, en Title Case. NO son socios ni administradores por sí mismos: buscá 'se designa apoderado', 'otorga poder a', 'representante legal', 'mandatario'. Si una misma persona es socio Y apoderado, incluila también acá. Si no hay apoderados: dejar vacío.",
  "dni_apoderados": "DNI de los apoderados, en el mismo orden que apoderados, separados por punto y coma, solo el número sin puntos (ej: 12345678). Si un apoderado no tiene DNI en el texto, dejá su posición vacía. Si no hay apoderados: dejar vacío.",
  "domicilio_sociedad": "Domicilio legal o sede social completo tal como figura en el texto (calle, número, piso, ciudad, provincia). Si no figura: dejar vacío.",
  "domicilio_calle_nro": "Solo la calle y el número (con piso/oficina si lo hay) del domicilio social, sin la localidad ni la provincia (ej: 'Aristides Villanueva 543 piso 3'). Extraído del mismo domicilio social. Si no figura: dejar vacío.",
  "domicilio_localidad": "Solo la localidad/ciudad o departamento del domicilio social (ej: 'Godoy Cruz', 'Ciudad de Mendoza', 'San Rafael'), sin la calle. Si no figura: dejar vacío.",
  "escribano": "Nombre completo del escribano/notario interviniente que autorizó la escritura o certificó las firmas del instrumento, en Title Case. Buscá 'ante mí', 'Escribano', 'Notario', 'por ante el escribano', 'firmas certificadas por'. Si no interviene escribano (o no figura): dejar vacío.",
  "registro_escribano": "Número de registro notarial del escribano y su jurisdicción tal como figura (ej: '23 - Ciudad de Mendoza' o 'Registro N° 145 de Godoy Cruz'). Si no figura: dejar vacío.",
  "email_sociedad": "Correo electrónico o domicilio electrónico constituido de la sociedad. Si no figura: N/A",
  "email_socios": "Correos electrónicos de los socios separados por punto y coma. Si no figuran: N/A"
}

Reglas importantes:
- El nombre NO debe incluir "CONSTITUCION DE", "INSCRIPCION DE", "MODIFICACION DE", ni similares.
- El nombre NO debe incluir comillas tipográficas ni rectas.
- Para tipo: usá "Fideicomiso" si el nombre o texto menciona "Fideicomiso". Usá "Union Transitoria" si menciona "Unión Transitoria", "U.T." o "UTE" como figura principal. Usá "Otro" solo si no encaja en ninguna de las categorías anteriores.
- Para tipo_acto: determiná el tipo según el encabezado o el verbo principal del bloque (CONSTITUCION = Constitución, AUMENTO DE CAPITAL = Aumento de capital, MODIFICACION = Modificación de estatuto, DISOLUCION = Disolución, LIQUIDACION = Liquidación, INSCRIPCION DE SUCURSAL = Inscripción de sucursal, TRANSFORMACION = Transformación, FUSION = Fusión, ESCISION = Escisión). Si no queda claro, usá "Constitución".
- Para capital en Fideicomisos: buscá "patrimonio fideicomitido" o "bienes fideicomitidos" — si figura un monto con $, extraélo igual. Si no figura monto, dejar vacío.
- Para capital en Uniones Transitorias: las UT generalmente no tienen capital social; dejar vacío salvo que figure explícitamente un monto.
- Para dni_socios: los números de DNI argentino no llevan puntos (12345678, no 12.345.678). Identificá el tipo de documento si no es DNI argentino.
- Para email_sociedad y email_socios: si no figuran en el texto, devolvé exactamente el string "N/A".
- Para capital: hay múltiples formatos posibles. Buscá CUALQUIERA de estos patrones y extraé el valor numérico:
  * "Capital de $5.000.000" — sin la palabra "social"
  * "capital: $200.000" — solo con dos puntos, sin "social"
  * "Capital: El capital se fija en la suma de $ 450.000,00" — descripción antes del monto
  * "capital social de PESOS CIEN MIL ($100.000)" — monto en palabras + número en paréntesis
  * "CAPITAL SOCIAL: ... PESOS UN MILLON ($1.500.000.-)" — número con punto y guión al final
  * "PESOS UN MILLON con 00/100 ($ 1.000.000)" — fracción "con 00/100" antes del paréntesis
  * "Capital Social: $300.000" o "monto del capital: $X"
  * "($600.000)" o "($ 500.000,00)" solos después de mencionar el capital
  Regla universal: si ves la palabra "capital" (sola o con "social") seguida de cualquier monto con $, ese es el capital.
  Ignorá menciones de "capital" que se refieran a la ciudad (ej: "departamento de Capital", "Ciudad Capital") o al objeto social (ej: "aportes de capitales a personas").
  Devolvé solo el símbolo $ y el número limpio, ej: $5.000.000 o $450.000,00. Eliminá el ".-" final si lo hubiera.
  CRÍTICO — No confundas capital total con valor nominal por acción/cuota: si el texto dice "$100 v/n c/u", "$100 valor nominal cada una", "$100 valor nominal por acción/cuota" o similar, ese monto es el precio por unidad, NO el capital. El capital es el monto total que aparece antes de esa frase (ej: en "Capital de $750.000 representado por 7500 acciones de $100 v/n c/u", el capital es $750.000).
- Para cargos_socios: los socios/accionistas se listan primero en el texto; luego aparece una cláusula de "ADMINISTRACIÓN" o "DIRECTORIO" o "GERENCIA" que nombra a los administradores/directores (que pueden o no ser los mismos socios). Si un socio también es nombrado administrador/director en esa cláusula, su cargo es "Socio / Administrador Titular" (o el que corresponda). Si alguien aparece SOLO en la cláusula de administración sin figurar antes como socio, incluyelo igualmente en nombres_socios con su cargo. En S.A.S. el "Administrador" equivale al Director; en S.R.L. se llama "Gerente"; en Asociación Civil es la "Comisión Directiva".
- Para domicilios_socios: el domicilio particular de cada socio aparece después de su nombre/datos personales y antes de los datos del siguiente socio. NO confundir con el domicilio social/sede de la empresa (que aparece en cláusulas como "SEDE SOCIAL:", "DOMICILIO SOCIAL:", "DOMICILIO LEGAL:"). Incluí calle, número, ciudad y provincia tal como figuran. El orden debe coincidir con nombres_socios; si un socio no tiene domicilio particular registrado usá posición vacía entre punto y coma.
- Para socios_juridicos y cuit_socios_juridicos: cuando un socio/accionista es otra sociedad (ej: "la firma XX S.A., CUIT 30-...", "la sociedad YY S.R.L.", "ZZ S.A.S. titular de N acciones"), NO la pongas en nombres_socios: va en socios_juridicos, con su CUIT en cuit_socios_juridicos (mismo orden). Mantené SEPARADAS las personas físicas (nombres_socios / dni_socios / cuit_socios / domicilios_socios / cargos_socios / profesiones_socios) de las personas jurídicas (socios_juridicos / cuit_socios_juridicos). Si una sociedad socia actúa "representada por" una persona física (su presidente, apoderado o quien firma), esa persona NO es socia: no la incluyas como socio salvo que el texto diga que también participa a título personal. En Uniones Transitorias, todos los integrantes van en socios_juridicos.
- Para porcentajes_socios y porcentajes_socios_juridicos: el porcentaje de cada socio debe quedar alineado por posición con nombres_socios y socios_juridicos respectivamente (mismo orden, separados por punto y coma, con posiciones vacías cuando falte). Si el texto expresa la participación en cantidad de acciones o cuotas (ej: 'suscribe 25 de las 50 acciones'), convertilo a porcentaje sobre el total (25/50 = 50%). Usá coma decimal para fracciones (ej: '33,33%'). No inventes porcentajes: si no figura ni la participación explícita ni datos suficientes para calcularla, dejá la posición vacía.
- Para apoderados y dni_apoderados: el apoderado/representante es alguien a quien la sociedad le otorga PODER para actuar en su nombre; NO confundir con socios, administradores ni gerentes (esos van en nombres_socios/cargos_socios). Una persona puede ser socio y además apoderado: en ese caso va en ambos campos. dni_apoderados debe quedar alineado por posición con apoderados.
- Para domicilio_calle_nro y domicilio_localidad: desglosá el MISMO domicilio social que ya pusiste en domicilio_sociedad — calle+número (y piso/oficina si lo hay) en uno, y la localidad/ciudad/departamento en el otro. No incluyas la provincia en domicilio_localidad salvo que sea lo único que figure. Deben ser coherentes con domicilio_sociedad.
- Para escribano y registro_escribano: en escrituras públicas figura el escribano que autoriza ('ante mí', 'por ante el escribano'); en instrumentos privados figura quien certifica las firmas. Capturá su nombre y, si está, el número de registro notarial con su jurisdicción. Si el acto no menciona escribano, dejá ambos vacíos.
- Para cuit_sociedad: es el CUIT de la PROPIA sociedad cuyo acto se publica, que suele aparecer en el encabezado de los actos que no son constitución (ej: 'NOMBRE S.A., CUIT 30-..., comunica...'). En las constituciones casi nunca figura: dejá vacío. No lo confundas con el CUIT de un socio.
- Para descripcion_acto: es el campo más útil para los actos que NO son constitución. Resumí el cambio concreto en 1-2 frases con nombres, montos y porcentajes (aumentos: de cuánto a cuánto; cesiones: quién cede cuántas cuotas/acciones a quién; designación de autoridades: qué cargos y quiénes; reformas: qué cláusula cambió). En constituciones podés dejarlo vacío porque los campos estructurados ya describen todo.
- Para fechas_nacimiento_socios: convertí cualquier formato a DD/MM/YYYY. Formatos frecuentes: "nacido el 15/03/1985", "nacida el 15 de marzo de 1985", "nacido el 15 de marzo de 1985", "de 38 años de edad, nacido el..." (en este caso extraé la fecha, no la edad). El orden debe coincidir exactamente con nombres_socios; si un socio no tiene fecha registrada en el texto usá una posición vacía entre punto y coma (ej: "15/03/1985; ; 22/11/1990").
- Si el texto es muy corto o no corresponde a una sociedad real, devolvé el JSON igualmente con los campos que puedas extraer (omitiendo el resto, como se explicó arriba).
- Formato de salida: un objeto JSON por bloque, en una sola línea, sin indentación, y SIN las claves de valor vacío (salvo email_sociedad/email_socios, que siempre van). Devolvé SOLO el/los JSON, nada más — sin numerarlos, sin arreglo, sin texto entre ellos.

---

## Ejemplos de extracción correcta

### Ejemplo 1 — S.R.L. con fecha numérica y dos socios

TEXTO:
CONSTITUCION DE SERVICIOS GARCIA Y ASOCIADOS S.R.L.- 1. SOCIOS: María Eugenia García Romero, argentina, comerciante, DNI N° 28.456.123, CUIT 27-28456123-3, con domicilio en calle Las Heras 456, Godoy Cruz, Mendoza; y Rodrigo Martín López Sánchez, argentino, contador público, DNI N° 32.789.456, CUIT 20-32789456-7, con domicilio en Av. San Martín 1234 piso 2, Ciudad de Mendoza. 2. DENOMINACION: Servicios García y Asociados S.R.L. 3. DOMICILIO SOCIAL: Calle Las Heras 456 piso 1 oficina A, Godoy Cruz, Mendoza. Domicilio electrónico constituido: serviciosgarcia@boe.mendoza.gov.ar. 4. OBJETO: Prestar servicios de consultoría contable, impositiva y financiera a personas humanas y jurídicas. 5. CAPITAL: Pesos doscientos cincuenta mil ($250.000), dividido en 250 cuotas de $1.000 cada una, suscriptas en partes iguales por ambos socios. 6. FECHA ACTO CONSTITUTIVO: 15/03/2023. 7. DURACION: 99 años. 8. APODERADO: Se designa apoderado al Sr. Juan Carlos Méndez, DNI 20.111.222, para realizar los trámites de inscripción.

JSON:
{"nombre": "Servicios Garcia Y Asociados S.R.L.", "tipo": "S.R.L.", "fecha_acto": "15/03/2023", "capital": "$250.000", "tipo_acto": "Constitución", "nombres_socios": "María Eugenia García Romero; Rodrigo Martín López Sánchez", "cargos_socios": "Socio / Gerente Titular; Socio / Gerente Suplente", "profesiones_socios": "comerciante; contador público", "objeto_social": "Prestar servicios de consultoría contable, impositiva y financiera a personas humanas y jurídicas.", "cuit_socios": "27284561233; 20327894567", "dni_socios": "28456123; 32789456", "domicilios_socios": "calle Las Heras 456, Godoy Cruz, Mendoza; Av. San Martín 1234 piso 2, Ciudad de Mendoza", "porcentajes_socios": "50%; 50%", "apoderados": "Juan Carlos Méndez", "dni_apoderados": "20111222", "domicilio_sociedad": "Calle Las Heras 456 piso 1 oficina A, Godoy Cruz, Mendoza", "domicilio_calle_nro": "Las Heras 456 piso 1 oficina A", "domicilio_localidad": "Godoy Cruz", "email_sociedad": "serviciosgarcia@boe.mendoza.gov.ar", "email_socios": "N/A"}

### Ejemplo 2 — S.A.S. con comillas tipográficas, emails de socios y fecha numérica

TEXTO:
CONSTITUCION DE "INNOVACION DIGITAL MENDOZA" S.A.S.- 1. SOCIOS FUNDADORES: Juan Pablo Herrera Torres, argentino, ingeniero en sistemas, DNI 25.678.901, CUIL 20-25678901-5, domicilio real Aristides Villanueva 543, Ciudad de Mendoza, correo electrónico jpherrera@gmail.com; Valentina Rocío Suárez Paz, argentina, diseñadora gráfica, DNI 33.445.678, CUIL 27-33445678-0, domicilio real Belgrano 789, Godoy Cruz, Mendoza, correo electrónico vsuarez@mail.com. 2. DENOMINACION: "Innovacion Digital Mendoza" S.A.S. 3. SEDE SOCIAL: Aristides Villanueva 543 piso 3, Ciudad de Mendoza. Domicilio electrónico: innovaciondigital@boe.mendoza.gov.ar. 4. OBJETO: Desarrollo de software, aplicaciones móviles, consultoría tecnológica y servicios digitales en general. 5. CAPITAL: Pesos quinientos mil ($ 500.000,00). 6. FECHA ACTO CONSTITUTIVO: 23/12/2022, instrumentado por escritura pública N° 87 ante el Escribano Roberto Díaz Funes, titular del Registro Notarial N° 56 de Ciudad de Mendoza.

JSON:
{"nombre": "Innovacion Digital Mendoza S.A.S.", "tipo": "S.A.S.", "fecha_acto": "23/12/2022", "capital": "$500.000,00", "tipo_acto": "Constitución", "nombres_socios": "Juan Pablo Herrera Torres; Valentina Rocío Suárez Paz", "cargos_socios": "Socio / Administrador Titular; Socio / Administrador Suplente", "profesiones_socios": "ingeniero en sistemas; diseñadora gráfica", "objeto_social": "Desarrollo de software, aplicaciones móviles, consultoría tecnológica y servicios digitales en general.", "cuit_socios": "20256789015; 27334456780", "dni_socios": "25678901; 33445678", "domicilios_socios": "Aristides Villanueva 543, Ciudad de Mendoza; Belgrano 789, Godoy Cruz, Mendoza", "domicilio_sociedad": "Aristides Villanueva 543 piso 3, Ciudad de Mendoza", "domicilio_calle_nro": "Aristides Villanueva 543 piso 3", "domicilio_localidad": "Ciudad de Mendoza", "escribano": "Roberto Díaz Funes", "registro_escribano": "56 - Ciudad de Mendoza", "email_sociedad": "innovaciondigital@boe.mendoza.gov.ar", "email_socios": "jpherrera@gmail.com; vsuarez@mail.com"}

### Ejemplo 3 — Asociación Civil con fecha en texto, email y cuatro miembros

TEXTO:
CONSTITUCION DE ASOCIACION CIVIL FUTURO JUSTO.- MIEMBROS FUNDADORES: 1) Ana Laura Rodríguez Medina, argentina, docente, DNI 22.567.890, CUIL 27-22567890-4, domicilio en Boulogne Sur Mer 2345, Las Heras, Mendoza. 2) Carlos Sebastián Pérez Villanueva, argentino, abogado, DNI 19.345.678, CUIL 20-19345678-2, domicilio en Rivadavia 890, Ciudad de Mendoza. 3) Marta Cecilia Gutiérrez Flores, argentina, médica, DNI 30.123.456, CUIL 27-30123456-1, domicilio en San Martín 456, Luján de Cuyo, Mendoza. 4) Roberto Alejandro Sánchez Torres, argentino, contador, DNI 26.789.012, CUIL 20-26789012-6, domicilio en Godoy Cruz 123, Maipú, Mendoza. DENOMINACION: Asociación Civil Futuro Justo. DOMICILIO: Boulogne Sur Mer 2345, Las Heras, Mendoza. Domicilio electrónico constituido: futurojusto@boe.mendoza.gov.ar. FECHA ACTO CONSTITUTIVO: 23 de diciembre de 2022. OBJETO: Promover la educación, cultura y el acceso a la justicia de sectores vulnerables de la sociedad mendocina mediante talleres, charlas y asistencia legal gratuita. CAPITAL INICIAL: Pesos cien mil ochocientos ($ 100.800).

JSON:
{"nombre": "Asociacion Civil Futuro Justo", "tipo": "Asociacion Civil", "fecha_acto": "23/12/2022", "capital": "$100.800", "tipo_acto": "Constitución", "nombres_socios": "Ana Laura Rodríguez Medina; Carlos Sebastián Pérez Villanueva; Marta Cecilia Gutiérrez Flores; Roberto Alejandro Sánchez Torres", "cargos_socios": "Socio Fundador / Presidente; Socio Fundador / Secretario; Socio Fundador / Tesorero; Socio Fundador", "profesiones_socios": "docente; abogado; médica; contador", "objeto_social": "Promover la educación, cultura y el acceso a la justicia de sectores vulnerables de la sociedad mendocina mediante talleres, charlas y asistencia legal gratuita.", "cuit_socios": "27225678904; 20193456782; 27301234561; 20267890126", "dni_socios": "22567890; 19345678; 30123456; 26789012", "domicilios_socios": "Boulogne Sur Mer 2345, Las Heras, Mendoza; Rivadavia 890, Ciudad de Mendoza; San Martín 456, Luján de Cuyo, Mendoza; Godoy Cruz 123, Maipú, Mendoza", "domicilio_sociedad": "Boulogne Sur Mer 2345, Las Heras, Mendoza", "email_sociedad": "futurojusto@boe.mendoza.gov.ar", "email_socios": "N/A"}

### Ejemplo 4 — Sucursal Extranjera con CI boliviana y fecha certificada

TEXTO:
INSCRIPCION DE SUCURSAL EXTRANJERA DE COMERCIAL MARAVILLAS DEL ALTIPLANO LIMITADA.- La sociedad fue constituida en el país de Bolivia, certificada el día 5 de diciembre de 2014. Representante legal en Argentina: Jorge Antonio Mamani Quispe, boliviano, comerciante, CI: 4523678, con domicilio en Dorrego 1890, Ciudad de Mendoza. Sede social en Argentina: Dorrego 1890, Ciudad de Mendoza, Provincia de Mendoza. CAPITAL asignado a la sucursal: $ 10.000.000. OBJETO: Importación, exportación y comercialización de productos alimenticios y artesanías del altiplano boliviano.

JSON:
{"nombre": "Comercial Maravillas Del Altiplano Limitada", "tipo": "Sucursal Extranjera", "fecha_acto": "05/12/2014", "capital": "$10.000.000", "tipo_acto": "Constitución", "nombres_socios": "Jorge Antonio Mamani Quispe", "cargos_socios": "Representante Legal", "profesiones_socios": "comerciante", "objeto_social": "Importación, exportación y comercialización de productos alimenticios y artesanías del altiplano boliviano.", "dni_socios": "CI: 4523678", "domicilios_socios": "Dorrego 1890, Ciudad de Mendoza", "domicilio_sociedad": "Dorrego 1890, Ciudad de Mendoza, Provincia de Mendoza", "email_sociedad": "N/A", "email_socios": "N/A"}

### Ejemplo 5 — Disolución y liquidación (estado especial, campos parciales)

TEXTO:
DISOLUCION Y LIQUIDACION DE TRANSPORTES ANDINOS S.A., CUIT 30-65432198-7.- Por Asamblea General Extraordinaria de fecha 10/05/2023 se resolvió la disolución anticipada y liquidación de la sociedad. Liquidador designado: Francisco Javier Moreno González, argentino, DNI 18.234.567, CUIT 20-18234567-9. Domicilio social: Acceso Este 4567, Lateral Sur, Guaymallén, Mendoza.

JSON:
{"nombre": "Transportes Andinos S.A.", "cuit_sociedad": "30654321987", "tipo": "S.A.", "fecha_acto": "10/05/2023", "tipo_acto": "Disolución", "descripcion_acto": "Se resolvió la disolución anticipada y liquidación de la sociedad; se designó liquidador a Francisco Javier Moreno González.", "nombres_socios": "Francisco Javier Moreno González", "cargos_socios": "Liquidador", "cuit_socios": "20182345679", "dni_socios": "18234567", "domicilio_sociedad": "Acceso Este 4567, Lateral Sur, Guaymallén, Mendoza", "email_sociedad": "N/A", "email_socios": "N/A"}

### Ejemplo 6 — Cuatro formatos alternativos de capital en un mismo bloque (casos reales)

TEXTO:
Caso A — "capital: $200.000" (solo con dos puntos, sin "social"):
plazo de duración: 99 años; capital: $200.000 representado por 1.000 acciones ordinarias nominativas no endosables.

Caso B — "Capital: El capital se fija en la suma de $ 450.000,00":
6°) Capital: El capital se fija en la suma de $ 450.000,00, representado en 100 acciones ordinarias, nominativas no endosables, de $4.500,00 cada una.

Caso C — "CAPITAL SOCIAL: ... ($1.500.000.-)":
CAPITAL SOCIAL: El Capital Social se fija en la suma de PESOS UN MILLON QUINIENTOS MIL ($1.500.000.-), representado por acciones ordinarias.

Caso D — "PESOS UN MILLON con 00/100 ($ 1.000.000)":
Capital Social: el capital social es de PESOS UN MILLON con 00/100 ($ 1.000.000), representado por un millón de acciones ordinarias nominativas no endosables de valor nominal un peso ($1).

Capital correcto para cada caso: A → $200.000 | B → $450.000,00 | C → $1.500.000 | D → $1.000.000

### Ejemplo 7 — Capital sin la palabra "social": formato "Capital de $X"

TEXTO:
TERRA CAPITAL SAS. Constitución: Instrumento Privado con firmas certificadas de fecha 24 de septiembre de 2025. 1.- Accionistas: Señor Ivan Mariano TORDI, DNI 28.247.071, CUIL 20-28247071-4, comerciante, con domicilio en Menedez 290, San Martín, Mendoza; y señor Carlos Silvestre AIROLDI, DNI 28.584.879, CUIL 20-28584879-3, comerciante, con domicilio en Av. Lima 31, San Martín, Mendoza. 2.- Denominación: TERRA CAPITAL SAS. 3.- Sede social: Menéndez 290, Ciudad, San Martín, Mendoza. 4.- Objeto: actividades comerciales, industriales y de servicios en general. 5.- Plazo de duración: 99 años. 6.- Capital de $5.000.000 representado por 50 acciones ordinarias, nominativas no endosables de $100.000 v/n c/u y de un voto, 100% suscriptas y 25% de integración. 7.- Administrador titular: Ivan Mariano TORDI, DNI 28.247.071.

JSON:
{"nombre": "Terra Capital Sas", "tipo": "S.A.S.", "fecha_acto": "24/09/2025", "capital": "$5.000.000", "tipo_acto": "Constitución", "nombres_socios": "Ivan Mariano Tordi; Carlos Silvestre Airoldi", "cargos_socios": "Socio / Administrador Titular; Socio / Administrador Suplente", "profesiones_socios": "comerciante; comerciante", "objeto_social": "Actividades comerciales, industriales y de servicios en general.", "cuit_socios": "20282470714; 20285848793", "dni_socios": "28247071; 28584879", "domicilios_socios": "Menedez 290, San Martín, Mendoza; Av. Lima 31, San Martín, Mendoza", "domicilio_sociedad": "Menéndez 290, Ciudad, San Martín, Mendoza", "email_sociedad": "N/A", "email_socios": "N/A"}

### Ejemplo 7 — Capital en palabras seguido de número entre paréntesis

TEXTO:
BEANT S.A.S. Constitución de sociedad anónima simplificada. 1°) Socios: URFALINO, JUAN FRANCO, DNI 30.674.477, CUIT 20-30674477-2, ingeniero, con domicilio en El Maitén 2415, Luján de Cuyo, Mendoza; y KARZOVNIK, ALAN IOAV, DNI 31.029.086, CUIT 20-31029086-7, comerciante, con domicilio en Granaderos 1240, Ciudad, Mendoza. 2°) Fecha del acto constitutivo: 16 de junio de 2023. 3°) Denominación: BEANT S.A.S. 4°) Domicilio: El Maitén 2415, Luján de Cuyo, Mendoza. 5°) Objeto Social: servicios de asesoramiento, consultoría y comercialización en general. 6°) Plazo: 99 años. 7°) Monto de Capital Social: Se suscribe totalmente el capital social de CIENTO OCHENTA MIL ($180.000) representado por 1800 acciones de pesos cien valor nominal cada una. 8°) Administrador Titular: URFALINO, JUAN FRANCO, DNI 30.674.477.

JSON:
{"nombre": "Beant S.A.S.", "tipo": "S.A.S.", "fecha_acto": "16/06/2023", "capital": "$180.000", "tipo_acto": "Constitución", "nombres_socios": "Juan Franco Urfalino; Alan Ioav Karzovnik", "cargos_socios": "Socio / Administrador Titular; Socio", "profesiones_socios": "ingeniero; comerciante", "objeto_social": "Servicios de asesoramiento, consultoría y comercialización en general.", "cuit_socios": "20306744772; 20310290867", "dni_socios": "30674477; 31029086", "domicilios_socios": "El Maitén 2415, Luján de Cuyo, Mendoza; Granaderos 1240, Ciudad, Mendoza", "domicilio_sociedad": "El Maitén 2415, Luján de Cuyo, Mendoza", "email_sociedad": "N/A", "email_socios": "N/A"}

### Ejemplo 8 — S.A.S. con fechas de nacimiento explícitas en el texto

TEXTO:
FOOTLONG S.A.S. Comunicase la constitución de una Sociedad por Acciones Simplificada conforme a las siguientes previsiones. 1°) Socios: Carlos Alberto Davalos Tello, DNI 94.105.201, CUIT 20-94105201-1, peruano, nacido el 01/08/1971, profesión: comerciante, estado civil: soltero, domiciliado en calle Granaderos N° 1717, Capital, Provincia de Mendoza; y María Soledad Ortiz Pereyra, argentina, DNI 30.456.789, CUIT 27-30456789-3, nacida el 15 de marzo de 1984, de profesión contadora pública, domiciliada en Av. San Martín 2340, Godoy Cruz, Mendoza. 2°) Denominación: FOOTLONG S.A.S. 3°) Domicilio: Las Heras 511, Capital. 4°) Objeto Social: gastronomía y servicios de alimentación en general. 5°) Capital: $500.000. 6°) Fecha del Acto Constitutivo: 13 de julio de 2023.

JSON:
{"nombre": "Footlong S.A.S.", "tipo": "S.A.S.", "fecha_acto": "13/07/2023", "capital": "$500.000", "tipo_acto": "Constitución", "nombres_socios": "Carlos Alberto Davalos Tello; María Soledad Ortiz Pereyra", "cargos_socios": "Socio; Socio", "fechas_nacimiento_socios": "01/08/1971; 15/03/1984", "profesiones_socios": "comerciante; contadora pública", "objeto_social": "Gastronomía y servicios de alimentación en general.", "cuit_socios": "20941052011; 27304567893", "dni_socios": "94105201; 30456789", "domicilios_socios": "calle Granaderos N° 1717, Capital, Mendoza; Av. San Martín 2340, Godoy Cruz, Mendoza", "domicilio_sociedad": "Las Heras 511, Capital, Mendoza", "email_sociedad": "N/A", "email_socios": "N/A"}

### Ejemplo 9 — S.A.S. con un socio persona física y un socio persona jurídica (otra sociedad)

TEXTO:
CONSTITUCION DE LOGISTICA ANDINA S.A.S.- 1. SOCIOS: Pablo Andrés Reynoso Díaz, argentino, comerciante, DNI 27.345.678, CUIT 20-27345678-9, con domicilio en San Lorenzo 234, Ciudad de Mendoza; y la sociedad TRANSPORTES DEL OESTE S.A., CUIT 30-71234567-8, con domicilio social en Ruta 7 Km 1042, Guaymallén, Mendoza, representada en este acto por su presidente Sr. Mario Gómez. 2. DENOMINACION: Logística Andina S.A.S. 3. SEDE SOCIAL: San Lorenzo 234, Ciudad de Mendoza. 4. OBJETO: transporte de cargas y logística integral. 5. CAPITAL: $2.000.000, suscripto en un 50% por cada uno de los dos socios. 6. FECHA ACTO CONSTITUTIVO: 10/04/2024. 7. Administrador titular: Pablo Andrés Reynoso Díaz.

JSON:
{"nombre": "Logistica Andina S.A.S.", "tipo": "S.A.S.", "fecha_acto": "10/04/2024", "capital": "$2.000.000", "tipo_acto": "Constitución", "nombres_socios": "Pablo Andrés Reynoso Díaz", "cargos_socios": "Socio / Administrador Titular", "profesiones_socios": "comerciante", "objeto_social": "Transporte de cargas y logística integral.", "cuit_socios": "20273456789", "dni_socios": "27345678", "domicilios_socios": "San Lorenzo 234, Ciudad de Mendoza", "porcentajes_socios": "50%", "socios_juridicos": "Transportes Del Oeste S.A.", "cuit_socios_juridicos": "30712345678", "porcentajes_socios_juridicos": "50%", "domicilio_sociedad": "San Lorenzo 234, Ciudad de Mendoza", "email_sociedad": "N/A", "email_socios": "N/A"}

### Ejemplo 10 — Unión Transitoria (los integrantes son sociedades, no personas físicas)

TEXTO:
CONSTITUCION DE VIAL ANDES S.A. - CONSTRUCCIONES DEL SUR S.A. UNION TRANSITORIA.- Las sociedades VIAL ANDES S.A., CUIT 30-70123456-7, y CONSTRUCCIONES DEL SUR S.A., CUIT 30-71987654-3, celebran el presente compromiso de Unión Transitoria para la ejecución de la obra "Repavimentación Ruta Provincial 89", con una participación del 60% para VIAL ANDES S.A. y del 40% para CONSTRUCCIONES DEL SUR S.A. Domicilio especial: Av. España 1200, Ciudad de Mendoza. Fecha del instrumento: 05/06/2024.

JSON:
{"nombre": "Vial Andes S.A. - Construcciones Del Sur S.A. Union Transitoria", "tipo": "Union Transitoria", "fecha_acto": "05/06/2024", "tipo_acto": "Constitución", "objeto_social": "Ejecución de la obra de repavimentación de la Ruta Provincial 89.", "socios_juridicos": "Vial Andes S.A.; Construcciones Del Sur S.A.", "cuit_socios_juridicos": "30701234567; 30719876543", "porcentajes_socios_juridicos": "60%; 40%", "domicilio_sociedad": "Av. España 1200, Ciudad de Mendoza", "email_sociedad": "N/A", "email_socios": "N/A"}

### Ejemplo 11 — Aumento de capital (acto posterior: trae CUIT de la sociedad, fecha de asamblea y descripción)

TEXTO:
ATENCIO REPUESTOS S.A.S., CUIT 30-71789456-2, comunica que por Reunión de Socios de fecha 22/10/2024 se resolvió aumentar el capital social de PESOS CIEN MIL ($100.000) a PESOS SEISCIENTOS MIL ($600.000). El socio Marcelo Atencio, DNI 27.111.333, suscribe la totalidad del aumento. Sede social: Tirasso 1450, Guaymallén, Mendoza.

JSON:
{"nombre": "Atencio Repuestos S.A.S.", "cuit_sociedad": "30717894562", "tipo": "S.A.S.", "fecha_acto": "22/10/2024", "capital": "$600.000", "tipo_acto": "Aumento de capital", "descripcion_acto": "Se aumentó el capital social de $100.000 a $600.000; el socio Marcelo Atencio suscribió la totalidad del aumento.", "nombres_socios": "Marcelo Atencio", "cargos_socios": "Socio", "dni_socios": "27111333", "domicilio_sociedad": "Tirasso 1450, Guaymallén, Mendoza", "domicilio_calle_nro": "Tirasso 1450", "domicilio_localidad": "Guaymallén", "email_sociedad": "N/A", "email_socios": "N/A"}"""


# ── Contadores globales de uso de la API ──────────────────────────────────────
_stats = {
    "llamadas":           0,
    "tokens_input":       0,   # tokens regulares (no cacheados)
    "tokens_output":      0,
    "tokens_cache_write": 0,   # tokens escritos al caché (cobrados a $0.25/M)
    "tokens_cache_read":  0,   # tokens leídos del caché (cobrados a $0.10/M)
    "truncados":          0,   # respuestas cortadas por max_tokens
    "reintentos":         0,   # llamadas repetidas por error recuperable
}

def _fmt_tiempo(segundos: float) -> str:
    """Formatea segundos como '1m 23s' o '45s'."""
    m, s = divmod(int(segundos), 60)
    return f"{m}m {s:02d}s" if m else f"{s}s"

def _costo_estimado() -> float:
    """Calcula el costo estimado en USD basado en los contadores.
    Precios Haiku 4.5 (USD/millón de tokens): input 1.00, output 5.00,
    cache write 1.25 (TTL 5m, el que usa este path en tiempo real), cache read 0.10."""
    s = _stats
    costo_input    = (s["tokens_input"]       / 1_000_000) * 1.00
    costo_c_write  = (s["tokens_cache_write"] / 1_000_000) * 1.25
    costo_c_read   = (s["tokens_cache_read"]  / 1_000_000) * 0.10
    costo_output   = (s["tokens_output"]      / 1_000_000) * 5.00
    return costo_input + costo_c_write + costo_c_read + costo_output


# ── Cliente Anthropic (singleton thread-safe) ─────────────────────────────────
import threading

_client: anthropic.Anthropic | None = None
_client_lock  = threading.Lock()
_stats_lock   = threading.Lock()
_chk_lock     = threading.Lock()

def get_client() -> anthropic.Anthropic:
    global _client
    with _client_lock:
        if _client is None:
            api_key = os.getenv("ANTHROPIC_API_KEY")
            if not api_key:
                raise ValueError(
                    "Variable de entorno ANTHROPIC_API_KEY no encontrada. "
                    "Ejecuta: set ANTHROPIC_API_KEY=tu_clave"
                )
            _client = anthropic.Anthropic(api_key=api_key)
        return _client


# ── Meses en español (para fallback de fecha) ─────────────────────────────────
MESES_ES = {
    "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
    "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
    "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12",
}


# ── Funciones auxiliares ───────────────────────────────────────────────────────

def extraer_fecha_boletin(nombre_archivo: str) -> str:
    """Extrae la fecha del nombre del archivo PDF (formato YYYYMMDD en el nombre)."""
    match = re.search(r"(\d{8})", nombre_archivo)
    if match:
        try:
            fecha_obj = datetime.strptime(match.group(1), "%Y%m%d")
            return fecha_obj.strftime("%d/%m/%Y")
        except Exception:
            return ""
    return ""


_IDS_MAPPING_FILE = Path(__file__).parent / "PDFs" / "boletines" / "ids_boletines.json"

def _cargar_mapping_ids() -> dict:
    """Carga el mapeo nombre_archivo -> id interno del boletín, generado por
    'Descargar boletines.py' (siempre en PDFs/boletines/ids_boletines.json,
    la ubicación canónica — NO depende de BOLETINES_DIR/BOLETIN_INPUT_DIR:
    si se procesa una carpeta con un subconjunto de PDFs (symlinks a
    PDFs/boletines/, ej. para un backfill parcial), esa carpeta no tiene su
    propia copia del mapeo, pero el nombre de archivo sigue siendo la clave
    correcta contra el mapeo canónico)."""
    if _IDS_MAPPING_FILE.exists():
        try:
            return json.loads(_IDS_MAPPING_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


_MAPPING_IDS = _cargar_mapping_ids()


def extraer_id_boletin(nombre_archivo: str) -> str:
    """Busca el id interno del boletín (el id de la URL verpdf/{id}) según el
    mapeo generado por el descargador. Vacío si el archivo no está mapeado."""
    return str(_MAPPING_IDS.get(nombre_archivo, ""))


def limpiar_texto(texto: str) -> str:
    """Colapsa espacios y saltos de línea en uno solo."""
    if not texto:
        return ""
    return re.sub(r"\s+", " ", texto).strip()


def limpiar_capital(valor: str) -> str:
    """
    Convierte el capital extraído por Claude a número limpio (sin $ ni puntos de miles).
    '$1.200.000,00' → '1200000.0'  |  '$500.000' → '500000.0'  |  '' → ''
    """
    if not valor or not valor.strip():
        return ""
    s = valor.strip().replace("$", "").strip()
    # Formato argentino: punto = miles, coma = decimal
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(".", "")
    s = re.sub(r"[\-\.]+$", "", s).strip()
    try:
        num = float(s)
        # Devolver entero si no tiene decimales significativos
        return str(int(num)) if num == int(num) else str(num)
    except ValueError:
        return valor  # Si no se puede parsear, devolver el original sin modificar


# Patrón de encabezados de página que pdfplumber inserta dentro del texto
_RE_ENCABEZADO_PAGINA = re.compile(
    r"Bolet[ií]n Oficial\s*[-–]\s*Gobierno de Mendoza[^\n]*\n?"
    r"(?:Ministerio[^\n]*\n?)?"
    r"(?:Subsecretar[^\n]*\n?)?",
    re.IGNORECASE,
)

_RE_IMPORTE_CERO = re.compile(r"Importe:\s*\$\s*0[^\n]*\n?", re.IGNORECASE)

# Pie administrativo de la publicación (costo del boletín, no dato de la sociedad),
# ej: "Boleto N°: ATM_5014962 Importe: $ 286 17/12/2020 (1 Pub.)" o "S/Cargo 23/04/2026
# (1 Pub.)". Siempre es lo ÚLTIMO del bloque (ancla $ = fin de bloque), así que nunca
# puede comerse contenido real: en el peor caso no matchea y el pie queda sin limpiar.
# Medido en muestra real: cubre ~94% de los bloques que tienen este pie.
_RE_PIE_PUBLICACION = re.compile(
    r"(?:S/Cargo\s*|(?:Boleto\s*N°?:?\s*[\w-]+\s*)?Importe:\s*\$\s*[\d.,]+\s*)?"
    r"\d{1,2}(?:-\d{1,2})*/\d{1,2}/\d{4}(?:-\d{1,2}/\d{1,2}/\d{4})?\s*"
    r"\(\d+\s*Pub\.?\)\.?\s*$",
    re.IGNORECASE,
)

def limpiar_encabezados(texto: str) -> str:
    """Elimina encabezados de página y artefactos de tabla que pdfplumber inserta en el texto."""
    texto = _RE_ENCABEZADO_PAGINA.sub(" ", texto)
    texto = _RE_IMPORTE_CERO.sub(" ", texto)
    texto = _RE_PIE_PUBLICACION.sub("", texto)
    return texto.strip()


def sanitizar_json(texto: str) -> str:
    """
    Fix A: Elimina saltos de línea y tabulaciones literales dentro de strings JSON.
    Claude a veces genera objeto_social con \\n reales (inválido en JSON).
    Ejemplo de error: 'Expecting , delimiter: line 8 column 27 (char 236)'
    """
    resultado = []
    dentro_string = False
    escape_next = False
    for char in texto:
        if escape_next:
            resultado.append(char)
            escape_next = False
        elif char == "\\" and dentro_string:
            resultado.append(char)
            escape_next = True
        elif char == '"':
            dentro_string = not dentro_string
            resultado.append(char)
        elif dentro_string and char in "\n\r\t":
            resultado.append(" ")
        else:
            resultado.append(char)
    return "".join(resultado)


def _extraer_jsons_multiples(texto: str) -> list[dict]:
    """
    Extrae TODOS los objetos JSON {...} presentes en un texto.
    Maneja el caso donde Claude devuelve dos o más sociedades en un solo bloque
    (error 'Extra data: line 15 column 4') recorriendo el texto con conteo de
    profundidad de llaves en lugar de una expresión regular greedy.
    """
    resultados = []
    i = 0
    while i < len(texto):
        if texto[i] != "{":
            i += 1
            continue
        # Encontrar la llave de cierre correspondiente
        depth = 0
        in_str = False
        esc = False
        j = i
        while j < len(texto):
            c = texto[j]
            if esc:
                esc = False
            elif c == "\\" and in_str:
                esc = True
            elif c == '"':
                in_str = not in_str
            elif not in_str:
                if c == "{":
                    depth += 1
                elif c == "}":
                    depth -= 1
                    if depth == 0:
                        fragmento = sanitizar_json(texto[i : j + 1])
                        try:
                            obj = json.loads(fragmento)
                            if isinstance(obj, dict):
                                resultados.append(obj)
                        except json.JSONDecodeError:
                            pass
                        i = j + 1
                        break
            j += 1
        else:
            # No se encontró cierre: saltar este '{'
            i += 1
    return resultados


_MESES_ES = {
    "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
    "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
    "septiembre": "09", "setiembre": "09", "octubre": "10",
    "noviembre": "11", "diciembre": "12",
}

def normalizar_fecha(valor: str) -> str:
    """
    Intenta convertir cualquier formato de fecha a DD/MM/YYYY.
    Si no puede parsearla, devuelve el original sin modificar.

    Formatos soportados:
      DD/MM/YYYY            → ya correcto
      D/M/YYYY              → padding a dos dígitos
      DD-MM-YYYY / DD.MM.YYYY
      YYYY-MM-DD (ISO)
      DD/MM/YY              → año de 2 dígitos
      15 de marzo de 2023   → texto con nombre de mes
      15 de marzo del 2023  → ídem con "del"
    """
    if not valor or not valor.strip():
        return ""
    s = valor.strip()

    # DD/MM/YYYY (ya correcto, solo validar padding)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"

    # DD-MM-YYYY o DD.MM.YYYY
    m = re.match(r"^(\d{1,2})[-.](\d{1,2})[-.](\d{4})$", s)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"

    # YYYY-MM-DD (ISO 8601)
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"

    # DD/MM/YY — año de 2 dígitos
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2})$", s)
    if m:
        anio = int(m.group(3))
        anio_completo = 2000 + anio if anio <= 50 else 1900 + anio
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{anio_completo}"

    # "15 de marzo de 2023" / "15 de marzo del 2023"
    m = re.match(
        r"^(\d{1,2})\s+de\s+([a-záéíóúñ]+)\s+de[l]?\s+(\d{4})$",
        s, re.IGNORECASE,
    )
    if m:
        mes = _MESES_ES.get(m.group(2).lower())
        if mes:
            return f"{int(m.group(1)):02d}/{mes}/{m.group(3)}"

    # Si no matchea ningún patrón, devolver el original
    return s


def normalizar_cuits(valor: str) -> str:
    """
    Elimina guiones y puntos de CUITs/CUILs. Maneja múltiples valores separados por ';'.
    '20-12345678-9; 27-87654321-3' → '20123456789; 27876543213'
    """
    if not valor or not valor.strip():
        return ""
    partes = [p.strip() for p in valor.split(";")]
    return "; ".join(re.sub(r"[-.]", "", p) for p in partes if p)


def normalizar_dnis(valor: str) -> str:
    """
    Elimina puntos de DNIs argentinos. Respeta prefijos de documentos extranjeros (CI:, PAS:, etc.).
    Maneja múltiples valores separados por ';'.
    '12.345.678; CI: 17623124-6; PAS: ABC123' → '12345678; CI: 17623124-6; PAS: ABC123'
    """
    if not valor or not valor.strip():
        return ""
    partes = [p.strip() for p in valor.split(";")]
    resultado = []
    for p in partes:
        if not p:
            continue
        # Documentos extranjeros con prefijo (CI:, PAS:, DNI:, LE:, LC:...) → intactos
        if re.match(r"^[A-Za-z]{2,3}\s*:", p):
            resultado.append(p)
        else:
            resultado.append(p.replace(".", ""))
    return "; ".join(resultado)


def _dict_a_sociedad(datos: dict, fecha_boletin: str, id_boletin: str = "") -> dict:
    """Convierte un dict devuelto por Claude en una fila con las columnas del Excel."""
    sociedad = {col: "" for col in COLUMNAS}
    sociedad["ID del boletín"]                          = id_boletin
    sociedad["Fecha publicación en boletín"]            = fecha_boletin
    sociedad["Nombre de la sociedad"]                   = str(datos.get("nombre", "")).strip()
    sociedad["CUIT de la sociedad"]                     = normalizar_cuits(str(datos.get("cuit_sociedad", "")))
    sociedad["Tipo de sociedad"]                        = str(datos.get("tipo", "")).strip()
    sociedad["Fecha del acto"]                   = normalizar_fecha(str(datos.get("fecha_acto", "")))
    sociedad["Capital inicial"]                         = limpiar_capital(str(datos.get("capital", "")))
    sociedad["Tipo de acto"]                            = str(datos.get("tipo_acto", "Constitución")).strip()
    sociedad["Descripción del acto"]                    = str(datos.get("descripcion_acto", "")).strip()
    sociedad["Objeto social"]                           = str(datos.get("objeto_social", "")).strip()
    sociedad["Nombres de los socios"]                   = str(datos.get("nombres_socios", "")).strip()
    sociedad["Cargos de los socios"]                    = str(datos.get("cargos_socios", "")).strip()
    sociedad["Fechas de nacimiento de los socios"]      = str(datos.get("fechas_nacimiento_socios", "")).strip()
    sociedad["Profesiones de los socios"]               = str(datos.get("profesiones_socios", "")).strip()
    sociedad["DNI de los socios"]                       = normalizar_dnis(str(datos.get("dni_socios", "")))
    sociedad["CUIT/CUIL de los socios"]                 = normalizar_cuits(str(datos.get("cuit_socios", "")))
    sociedad["Domicilios particulares de los socios"]   = str(datos.get("domicilios_socios", "")).strip()
    sociedad["Porcentaje de los socios"]                = str(datos.get("porcentajes_socios", "")).strip()
    sociedad["Nombres de los socios jurídicos"]         = str(datos.get("socios_juridicos", "")).strip()
    sociedad["CUIT de los socios jurídicos"]            = normalizar_cuits(str(datos.get("cuit_socios_juridicos", "")))
    sociedad["Porcentaje de los socios jurídicos"]      = str(datos.get("porcentajes_socios_juridicos", "")).strip()
    sociedad["Nombres de los apoderados"]               = str(datos.get("apoderados", "")).strip()
    sociedad["DNI de los apoderados"]                   = normalizar_dnis(str(datos.get("dni_apoderados", "")))
    sociedad["Domicilio de la sociedad"]                = str(datos.get("domicilio_sociedad", "")).strip()
    sociedad["Calle y número de la sociedad"]           = str(datos.get("domicilio_calle_nro", "")).strip()
    sociedad["Localidad de la sociedad"]                = str(datos.get("domicilio_localidad", "")).strip()
    sociedad["Escribano interviniente"]                 = str(datos.get("escribano", "")).strip()
    sociedad["Registro notarial"]                       = str(datos.get("registro_escribano", "")).strip()
    sociedad["Domicilio electrónico de la sociedad"]    = str(datos.get("email_sociedad", "N/A")).strip() or "N/A"
    sociedad["Domicilio electrónico de los socios"]     = str(datos.get("email_socios", "N/A")).strip() or "N/A"
    return sociedad


# ── Extracción de datos vía Claude API ────────────────────────────────────────

def _construir_mensaje_usuario(bloques: list[str]) -> str:
    """Arma el texto del mensaje de usuario a partir de 1+ bloques ya limpios.

    Con 1 bloque: mismo mensaje de siempre (no cambia nada para quien ya
    procesaba de a uno). Con 2+: los concatena con un separador "=== BLOQUE N
    ===" — el SYSTEM_PROMPT ya sabe manejar esta forma y devuelve un JSON por
    bloque. Agrupar bloques reduce el número de llamadas (y por lo tanto
    cuántas veces se paga la lectura del caché del system prompt)."""
    if len(bloques) == 1:
        return f"Extraé los datos de esta sociedad:\n\n{bloques[0][:12000]}"
    partes = "\n\n".join(f"=== BLOQUE {i} ===\n{b[:12000]}" for i, b in enumerate(bloques, start=1))
    return (
        f"A continuación hay {len(bloques)} bloques independientes, cada uno de una "
        f"sociedad/entidad distinta. Extraé los datos de CADA bloque por separado y "
        f"devolvé un JSON por bloque, en el mismo orden:\n\n{partes}"
    )


def _construir_params_mensaje(bloques: str | list[str], cache_ttl: str | None = None) -> dict:
    """Construye los kwargs de la llamada a la API. Compartido por el path en
    tiempo real (messages.create) y el path de Batch API (params del request),
    para garantizar que ambos envían exactamente el mismo prompt.

    bloques: un bloque (str) o una lista de bloques a agrupar en una sola llamada.

    cache_ttl: vida del caché del system prompt.
        None -> caché efímero default (5 min). Es lo que conviene en tiempo real,
                donde las llamadas son seguidas y reusan el caché naturalmente.
        '1h' -> caché de 1 hora. Lo usa el path Batch (pre-calentado): así el
                caché sobrevive toda la corrida del lote y los requests lo LEEN
                en vez de reescribirlo en cada uno (que es lo caro)."""
    if isinstance(bloques, str):
        bloques = [bloques]
    cache_control: dict = {"type": "ephemeral"}
    if cache_ttl:
        cache_control["ttl"] = cache_ttl
    return {
        "model": MODEL,
        "max_tokens": MAX_TOKENS,
        "system": [
            {
                "type": "text",
                "text": SYSTEM_PROMPT,
                "cache_control": cache_control,
            }
        ],
        "messages": [
            {
                "role": "user",
                "content": _construir_mensaje_usuario(bloques),
            }
        ],
    }


def _parsear_texto_a_sociedades(texto_respuesta: str, fecha_boletin: str, id_boletin: str = "") -> list[dict]:
    """Convierte el texto crudo devuelto por el modelo en filas (dicts con COLUMNAS).
    Devuelve [] si no hay JSON válido. Compartido por ambos paths."""
    texto_respuesta = re.sub(r"```(?:json)?\s*", "", (texto_respuesta or "").strip()).strip()
    lista_datos = _extraer_jsons_multiples(texto_respuesta)
    if not lista_datos:
        return []
    return [_dict_a_sociedad(d, fecha_boletin, id_boletin) for d in lista_datos]


def _es_error_de_creditos(e: Exception) -> bool:
    """Detecta 'sin créditos' aunque llegue como 400 (BadRequestError) en vez del
    403 (PermissionDeniedError) esperable. Mira tanto str(e) como e.body: el
    formato exacto de str(e) depende de cómo el SDK arma el mensaje, así que
    e.body (el JSON decodificado de la respuesta) es la fuente más confiable."""
    texto = str(e)
    body = getattr(e, "body", None)
    if isinstance(body, dict):
        texto += " " + str(body.get("error", {}).get("message", ""))
    return "credit balance" in texto.lower()


def extraer_datos_claude(bloques: list[str], fecha_boletin: str, id_boletin: str = "") -> list[dict]:
    """
    Envía uno o más bloques (ya limpios y filtrados por el llamador, ver
    procesar_pdf) a Claude Haiku EN UNA SOLA LLAMADA y parsea la respuesta.
    Agrupar bloques (BLOQUES_POR_LLAMADA) reduce cuántas veces se paga la
    lectura del caché del system prompt por boletín.

    Retorna una LISTA de dicts — normalmente len(bloques) elementos, pero
    puede haber más (un bloque describía varias sociedades) o menos (algún
    bloque no arrojó datos).
    """
    vacia = [{col: "" for col in COLUMNAS}]
    vacia[0]["ID del boletín"] = id_boletin
    vacia[0]["Fecha publicación en boletín"] = fecha_boletin

    client = get_client()

    for intento in range(1, MAX_REINTENTOS + 1):
        t0_llamada = time.time()
        try:
            response = client.messages.create(**_construir_params_mensaje(bloques))

            elapsed = time.time() - t0_llamada
            u = response.usage
            with _stats_lock:
                _stats["llamadas"]           += 1
                _stats["tokens_input"]       += u.input_tokens
                _stats["tokens_output"]      += u.output_tokens
                _stats["tokens_cache_write"] += getattr(u, "cache_creation_input_tokens", 0) or 0
                _stats["tokens_cache_read"]  += getattr(u, "cache_read_input_tokens", 0)     or 0

            cache_tag = ""
            if getattr(u, "cache_read_input_tokens", 0):
                cache_tag = " [cache HIT]"
            elif getattr(u, "cache_creation_input_tokens", 0):
                cache_tag = " [cache WRITE]"

            tag_intento = f" [reintento {intento}/{MAX_REINTENTOS}]" if intento > 1 else ""
            if response.stop_reason == "max_tokens":
                _stats["truncados"] += 1
                logging.warning("      [!] Respuesta TRUNCADA (max_tokens). Considerar subir MAX_TOKENS.")

            logging.info(
                f"      API: {elapsed:.1f}s | "
                f"in={u.input_tokens} out={u.output_tokens} | "
                f"stop={response.stop_reason}{cache_tag}{tag_intento}"
            )

            texto_respuesta = response.content[0].text.strip()
            sociedades = _parsear_texto_a_sociedades(texto_respuesta, fecha_boletin, id_boletin)

            if not sociedades:
                logging.warning("      Respuesta de Claude sin JSON válido.")
                logging.warning(f"      Respuesta recibida: {texto_respuesta[:300]!r}")
                return vacia

            if len(sociedades) != len(bloques):
                logging.info(
                    f"      [{len(bloques)} bloque(s) enviados, {len(sociedades)} sociedad(es) devueltas]"
                )

            return sociedades

        except anthropic.RateLimitError:
            espera = 30 * intento   # 30s, 60s, 90s
            if intento < MAX_REINTENTOS:
                logging.warning(f"      Rate limit (intento {intento}/{MAX_REINTENTOS}). Esperando {espera}s...")
                with _stats_lock:
                    _stats["reintentos"] += 1
                time.sleep(espera)
            else:
                logging.error(f"      Rate limit agotado tras {MAX_REINTENTOS} intentos.")

        except (anthropic.APIConnectionError, anthropic.APITimeoutError) as e:
            espera = ESPERA_BASE * (2 ** (intento - 1))   # 5s, 10s, 20s
            if intento < MAX_REINTENTOS:
                logging.warning(f"      Error de conexión (intento {intento}/{MAX_REINTENTOS}): {e}. Esperando {espera}s...")
                with _stats_lock:
                    _stats["reintentos"] += 1
                time.sleep(espera)
            else:
                logging.error(f"      Error de conexión agotado tras {MAX_REINTENTOS} intentos: {e}")

        except anthropic.PermissionDeniedError as e:
            # Créditos agotados — propagar para que el PDF NO quede marcado en el checkpoint
            logging.error(f"\n{'!'*60}")
            logging.error(f"      SIN CRÉDITOS API (o error de permisos): {e}")
            logging.error(f"      ► Recargá créditos en console.anthropic.com y volvé a correr.")
            logging.error(f"      ► El checkpoint preserva los PDFs ya completados.")
            logging.error(f"{'!'*60}\n")
            raise   # No atrapar — el PDF no se sumará al checkpoint

        except anthropic.BadRequestError as e:
            # ARCA a veces manda "sin créditos" como 400 (invalid_request_error),
            # no como 403 (PermissionDeniedError) — mismo tratamiento: propagar sin
            # marcar el PDF, para no dejarlo "procesado" con datos vacíos/incompletos
            # que después se saltearían para siempre en un resume.
            if _es_error_de_creditos(e):
                logging.error(f"\n{'!'*60}")
                logging.error(f"      SIN CRÉDITOS API (400): {e}")
                logging.error(f"      ► Recargá créditos en console.anthropic.com y volvé a correr.")
                logging.error(f"      ► El checkpoint preserva los PDFs ya completados.")
                logging.error(f"{'!'*60}\n")
                raise
            logging.error(f"      Error 400 (request inválido) en llamada a Claude: {e}")
            break   # no es de créditos: no tiene sentido reintentar el mismo request

        except Exception as e:
            logging.error(f"      Error no recuperable en llamada a Claude: {e}")
            break   # No tiene sentido reintentar

    return vacia


# ── Procesamiento de PDF ───────────────────────────────────────────────────────

def extraer_bloques_seccion(ruta_pdf: str) -> list[str]:
    """
    Lee el PDF, detecta la sección CONTRATOS SOCIALES y devuelve la lista de
    bloques (*) crudos. Devuelve [] si no hay sección o no se puede leer.
    Compartido por el path en tiempo real y el de Batch API.
    """
    try:
        partes_texto = []
        paginas_rotas = []
        with pdfplumber.open(ruta_pdf) as pdf:
            for nro, page in enumerate(pdf.pages, start=1):
                # Página por página con try/except: algunos boletines traen UNA
                # página con el content-stream mal formado que hace que pdfminer
                # tire "'int'/'float'/'PSLiteral' object is not iterable". Si se
                # extrae todo de una sola vez, esa única página corrupta aborta el
                # boletín ENTERO y se pierden todas sus sociedades (pasó con 12
                # boletines de 2019-2021). Aislando cada página salteamos solo la
                # rota y conservamos el resto (la sección CONTRATOS SOCIALES casi
                # siempre está en otra página y se recupera intacta).
                try:
                    partes_texto.append(page.extract_text() or "")
                except Exception as e_pag:
                    paginas_rotas.append(nro)
                    partes_texto.append("")
        # "\n" (no ""): page.extract_text() no siempre termina en salto de línea,
        # así que unir sin separador pega la última línea de una página con la
        # primera de la siguiente sin espacio. Eso confundía a limpiar_encabezados
        # (el regex de encabezado de página terminaba comiéndose contenido real
        # del bloque siguiente al no encontrar un \n donde esperaba cortar).
        texto = "\n".join(partes_texto)
        if paginas_rotas:
            logging.warning(
                f"  {Path(ruta_pdf).name}: {len(paginas_rotas)} página(s) "
                f"ilegibles salteadas (pág. {paginas_rotas}); se procesa el resto."
            )
    except Exception as e:
        logging.error(f"Error leyendo {Path(ruta_pdf).name}: {e}")
        return []

    if "CONTRATOS SOCIALES" not in texto:
        logging.info(f"  Sin seccion CONTRATOS SOCIALES en {Path(ruta_pdf).name}")
        return []

    # "CONTRATOS SOCIALES" aparece primero en el índice (con puntos suspensivos y
    # número de página) y de nuevo en el título real de la sección. Si se toma la
    # PRIMERA aparición (el índice) como punto de partida para buscar el primer
    # "(*)", cualquier "(*)" de OTRA sección que aparezca ANTES del título real
    # (ej: notas al pie de tablas de tasas, cuotas, zonificación — usan "(*)"
    # como marca de nota al pie en todo el boletín, no solo acá) hace arrancar
    # la extracción en un lugar totalmente ajeno. Hay que ubicar el título REAL,
    # saltando cualquier mención que sea claramente la entrada del índice.
    pos = 0
    titulo_real = -1
    while True:
        pos = texto.find("CONTRATOS SOCIALES", pos)
        if pos == -1:
            break
        resto = texto[pos + len("CONTRATOS SOCIALES"): pos + len("CONTRATOS SOCIALES") + 20]
        if re.match(r"\s*\.{5,}", resto):   # entrada de índice: le siguen puntos suspensivos
            pos += 1
            continue
        titulo_real = pos
        break
    if titulo_real == -1:
        return []

    primer_asterisco = texto.find("(*)", titulo_real)
    if primer_asterisco == -1:
        return []

    # Buscar fin de sección (próxima sección del boletín). Regex con \n al inicio
    # para que solo matchee títulos de sección reales (al comienzo de una línea).
    fin_seccion = len(texto)
    subtexto = texto[primer_asterisco:]
    for seccion_siguiente in SECCIONES_SIGUIENTES:
        m = re.search(r"\n\s*" + re.escape(seccion_siguiente) + r"[\s\n]", subtexto)
        if m:
            idx = primer_asterisco + m.start()
            if idx < fin_seccion:
                fin_seccion = idx

    seccion = texto[primer_asterisco:fin_seccion]
    return re.split(r"\(\*\)", seccion)[1:]  # [0] es vacío antes del primer (*)


def procesar_pdf(ruta_pdf: str, fecha_boletin: str, id_boletin: str = "") -> list[dict]:
    """
    Procesa un PDF en tiempo real: extrae los bloques de la sección
    CONTRATOS SOCIALES, los agrupa de a BLOQUES_POR_LLAMADA y llama a Claude
    una vez por grupo (en vez de una vez por bloque).
    """
    try:
        bloques_crudos = extraer_bloques_seccion(ruta_pdf)
        if not bloques_crudos:
            return []

        # Limpiar y filtrar por longitud ANTES de agrupar (mismo criterio que
        # extraer_batch.py: bloques < 50 chars no se envían).
        bloques = [
            limpio
            for b in bloques_crudos
            if len(limpio := limpiar_texto(limpiar_encabezados(b))) >= 50
        ]
        if not bloques:
            return []

        grupos = [bloques[i:i + BLOQUES_POR_LLAMADA] for i in range(0, len(bloques), BLOQUES_POR_LLAMADA)]
        logging.info(
            f"  {len(bloques)} bloque(s) encontrado(s) en {Path(ruta_pdf).name} "
            f"→ {len(grupos)} llamada(s) (de a {BLOQUES_POR_LLAMADA})"
        )

        t0_pdf = time.time()
        # Los grupos de un mismo PDF son independientes entre sí (cada llamada a
        # Claude no depende de la respuesta de otra) — se paralelizan con un pool
        # chico en vez de mandarlos secuencialmente. Con 1 solo PDF/día (caso típico
        # del job diario) el ThreadPoolExecutor de más abajo (por ARCHIVO) no ayuda
        # en nada; este sí, porque un boletín con 7 llamadas de ~9s c/u pasa de ~60s
        # secuenciales a ~2 tandas en paralelo.
        from concurrent.futures import ThreadPoolExecutor as _TPE, as_completed as _as_completed
        resultados_por_grupo: list[list[dict]] = [[] for _ in grupos]
        n_workers = min(BLOQUES_EN_PARALELO, len(grupos))
        with _TPE(max_workers=n_workers) as pool:
            futuros = {
                pool.submit(extraer_datos_claude, grupo, fecha_boletin, id_boletin): i
                for i, grupo in enumerate(grupos)
            }
            completados = 0
            for futuro in _as_completed(futuros):
                i = futuros[futuro]
                resultados_por_grupo[i] = futuro.result()
                completados += 1
                logging.info(f"    Llamada {completados}/{len(grupos)} ({len(grupos[i])} bloque(s)) OK")

        sociedades = [
            sociedad
            for resultado in resultados_por_grupo
            for sociedad in resultado
            if sociedad.get("Nombre de la sociedad")
        ]

        logging.info(f"  PDF procesado en {_fmt_tiempo(time.time() - t0_pdf)} — {len(sociedades)} sociedad(es)")
        return sociedades

    except (anthropic.PermissionDeniedError, anthropic.BadRequestError):
        # BadRequestError llega hasta acá SOLO cuando extraer_datos_claude ya
        # confirmó que es el caso de "sin créditos" (400) y decidió relanzarlo
        # — los 400 de otro tipo se resuelven ahí mismo y no propagan.
        raise   # Propagar: el loop principal no marcará este PDF en el checkpoint

    except Exception as e:
        logging.error(f"Error procesando {Path(ruta_pdf).name}: {e}")
        return []


# ── Detección de duplicados ───────────────────────────────────────────────────

# Columnas que definen si dos filas son "el mismo acto" (100% idénticas).
# Se excluyen "ID del boletín" y "Fecha publicación en boletín" a propósito:
# así se detecta el mismo acto aunque haya salido publicado por error en dos
# boletines distintos, pero NO se marca como duplicado un segundo acto
# legítimo del mismo tipo para la misma sociedad (ej. dos "Aumento de
# capital" en años distintos, con capital/fecha/socios distintos).
_COLUMNAS_CLAVE_DUPLICADO = [
    c for c in COLUMNAS if c not in ("ID del boletín", "Fecha publicación en boletín")
]


def _marcar_duplicados(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega la columna 'Duplicado' al DataFrame.

    Lógica:
      - Clave: todo el contenido extraído (ver _COLUMNAS_CLAVE_DUPLICADO) —
        dos filas son duplicado solo si son el mismo acto, 100% idéntico.
      - La primera aparición (orden cronológico por fecha de publicación) = 'No'
      - Apariciones posteriores con contenido idéntico = 'Sí'

    El DataFrame ya viene ordenado por fecha de publicación, por lo que
    'primera aparición' equivale siempre al registro más antiguo.
    """
    clave = df[_COLUMNAS_CLAVE_DUPLICADO].fillna("").astype(str).agg("§".join, axis=1)

    es_duplicado = clave.duplicated(keep="first").map({True: "Sí", False: "No"})

    pos = df.columns.get_loc("Tipo de acto") + 1
    df.insert(pos, "Duplicado", es_duplicado)
    return df


# ── Checkpoint ────────────────────────────────────────────────────────────────

def _cargar_checkpoint() -> tuple[set, list]:
    """Carga el checkpoint si existe. Devuelve (archivos_ya_procesados, sociedades_acumuladas)."""
    if not CHECKPOINT_FILE.exists():
        return set(), []
    try:
        with open(CHECKPOINT_FILE, encoding="utf-8") as f:
            data = json.load(f)
        procesados = set(data.get("procesados", []))
        sociedades = data.get("sociedades", [])
        logging.info(
            f"Checkpoint encontrado: {len(procesados)} PDF(s) ya procesados, "
            f"{len(sociedades)} sociedad(es) acumuladas."
        )
        return procesados, sociedades
    except Exception as e:
        logging.warning(f"No se pudo leer el checkpoint ({e}). Empezando desde cero.")
        return set(), []


def _guardar_checkpoint(procesados: set, sociedades: list) -> None:
    """Persiste el estado actual al archivo de checkpoint.

    Escritura atómica (a un .tmp + rename), igual que ya hace el descargador
    de PDFs: si el proceso se corta a mitad de la escritura (Ctrl+C, cierre de
    terminal, falla de energía), el .tmp queda a medio escribir pero el
    checkpoint real en disco NO se toca hasta que el rename confirma que la
    escritura terminó completa. Así nunca queda un checkpoint corrupto."""
    tmp = CHECKPOINT_FILE.with_suffix(CHECKPOINT_FILE.suffix + ".tmp")
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump({"procesados": sorted(procesados), "sociedades": sociedades},
                      f, ensure_ascii=False)
        tmp.replace(CHECKPOINT_FILE)   # rename atómico al terminar
    except Exception as e:
        logging.warning(f"No se pudo guardar el checkpoint: {e}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    t0_total = time.time()
    carpeta = BOLETINES_PRUEBA if TEST_MODE else BOLETINES_DIR

    if not carpeta.exists():
        logging.error(f"Carpeta no encontrada: {carpeta}")
        return

    archivos = sorted(carpeta.glob("*.pdf"))
    if not archivos:
        logging.error(f"No hay archivos PDF en {carpeta.name}/")
        return

    if TEST_MODE:
        logging.info(f"[MODO PRUEBA] Carpeta: {carpeta.name} — {len(archivos)} archivos")
    else:
        logging.info(f"Procesando {len(archivos)} archivos PDF...")

    procesados, todas_las_sociedades = _cargar_checkpoint()
    pendientes = [a for a in archivos if a.name not in procesados]
    errores = 0

    if procesados:
        logging.info(f"Saltando {len(procesados)} PDF(s) ya procesados. Quedan {len(pendientes)}.")

    modo = f"{MAX_WORKERS} workers en paralelo" if MAX_WORKERS > 1 else "modo secuencial"
    logging.info(f"Iniciando procesamiento ({modo})...")

    from concurrent.futures import ThreadPoolExecutor, CancelledError

    def _procesar_uno(archivo):
        fecha = extraer_fecha_boletin(archivo.name)
        id_boletin = extraer_id_boletin(archivo.name)
        return archivo, procesar_pdf(str(archivo), fecha, id_boletin)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futuros = {executor.submit(_procesar_uno, arch): arch for arch in pendientes}
        restantes = dict(futuros)   # se van sacando a medida que se procesan
        completados = len(procesados)
        interrumpido = False
        sin_creditos = False

        def _procesar_uno_listo(futuro) -> None:
            nonlocal completados, errores, sin_creditos
            completados += 1
            try:
                archivo, sociedades = futuro.result()
                todas_las_sociedades.extend(sociedades)
                with _chk_lock:
                    procesados.add(archivo.name)
                    _guardar_checkpoint(procesados, todas_las_sociedades)
                logging.info(f"  [{completados}/{len(archivos)}] {archivo.name} → {len(sociedades)} sociedad(es) [checkpoint guardado]")
            except CancelledError:
                pass   # no había arrancado todavía: no cuenta como error ni se procesa
            except (anthropic.PermissionDeniedError, anthropic.BadRequestError):
                # Llega hasta acá ya confirmado que es "sin créditos" (ver procesar_pdf
                # y extraer_datos_claude) — no se marca procesado, y le avisamos al
                # bucle principal para que corte TODA la corrida, no solo este PDF.
                archivo = futuros[futuro]
                logging.error(f"  [{completados}/{len(archivos)}] {archivo.name} → sin créditos, no se marca como procesado")
                errores += 1
                sin_creditos = True
            except Exception as e:
                archivo = futuros[futuro]
                logging.error(f"  [{completados}/{len(archivos)}] {archivo.name} → Error fatal: {e}")
                errores += 1

        def _cancelar_y_drenar(motivo: str) -> None:
            logging.warning(
                f"[!] {motivo}: cancelando los PDFs que todavía no arrancaron a "
                "procesarse. Esperando a que terminen los que ya estaban en curso..."
            )
            executor.shutdown(wait=False, cancel_futures=True)
            while restantes:
                listos = [f for f in restantes if f.done()]
                for f in listos:
                    _procesar_uno_listo(f)
                    del restantes[f]
                if restantes:
                    time.sleep(0.2)

        # Todos los PDFs pendientes se mandan al pool de una sola vez (arriba), así
        # que hay que sondear los resultados a mano en vez de "for f in as_completed(...)":
        # esperar sin límite de tiempo dentro de as_completed()/concurrent.futures.wait()
        # puede TRAGARSE la señal de Ctrl+C (probado: no es confiable — a veces la
        # pierde del todo y el proceso sigue de largo sin enterarse). Con .done() +
        # time.sleep() cortos, el intérprete tiene chances frecuentes de notar la
        # señal, y time.sleep() sí es interrumpible de forma confiable.
        try:
            while restantes:
                listos = [f for f in restantes if f.done()]
                for f in listos:
                    _procesar_uno_listo(f)
                    del restantes[f]
                if sin_creditos:
                    _cancelar_y_drenar("Sin créditos API")
                    logging.info(
                        f"Corrida detenida por falta de créditos. {len(procesados)} PDF(s) "
                        "quedaron guardados en el checkpoint. Recargá créditos en "
                        "console.anthropic.com y volvé a correr el mismo comando."
                    )
                    return
                if restantes:
                    time.sleep(0.2)
        except KeyboardInterrupt:
            interrumpido = True
            _cancelar_y_drenar("Ctrl+C recibido")
            logging.info(f"Corrida detenida a pedido (Ctrl+C). {len(procesados)} PDF(s) quedaron guardados en el checkpoint.")
            return

    logging.info(f"\nTotal: {len(todas_las_sociedades)} sociedades en {len(archivos) - errores}/{len(archivos)} archivos.")

    s = _stats
    logging.info(
        f"Uso API: {s['llamadas']} llamadas | "
        f"tokens in={s['tokens_input']:,} out={s['tokens_output']:,} | "
        f"cache write={s['tokens_cache_write']:,} read={s['tokens_cache_read']:,} | "
        f"truncados={s['truncados']} | "
        f"reintentos={s['reintentos']} | "
        f"costo estimado=${_costo_estimado():.4f}"
    )

    if not todas_las_sociedades:
        print("\n[ADVERTENCIA] No se encontraron sociedades.")
        return

    # ── DataFrame y Excel ────────────────────────────────────────────────────
    df = pd.DataFrame(todas_las_sociedades, columns=COLUMNAS)
    df = df[df["Nombre de la sociedad"].str.strip() != ""]
    # Orden CRONOLÓGICO real: "Fecha publicación en boletín" es texto "DD/MM/YYYY",
    # ordenar por ese string directamente compara primero el día ("01/02/2018"
    # queda antes que "01/03/2017") — hay que ordenar por la fecha real, no el texto.
    _orden = pd.to_datetime(df["Fecha publicación en boletín"], format="%d/%m/%Y", errors="coerce")
    df = (df.assign(_orden=_orden)
            .sort_values(["_orden", "Nombre de la sociedad"])
            .drop(columns="_orden")
            .reset_index(drop=True))

    df = _marcar_duplicados(df)
    n_duplicados = (df["Duplicado"] == "Sí").sum()
    logging.info(f"Duplicados detectados: {n_duplicados} de {len(df)} registros.")

    # Preview en consola
    print(f"\n{'='*100}")
    print(f"RESULTADO: {len(df)} sociedad(es) extraida(s)")
    print(f"{'='*100}")
    for idx, row in df.iterrows():
        print(f"\n[{idx+1}] {row['Nombre de la sociedad']}")
        print(f"    Tipo:    {row['Tipo de sociedad']}")

        print(f"    Fecha:   {row['Fecha del acto']}")
        print(f"    Capital: {row['Capital inicial']}")
        print(f"    Acto:    {row['Tipo de acto']}")
        if row["Nombres de los socios"]:
            print(f"    Nombres: {row['Nombres de los socios'][:100]}")
        if row["Profesiones de los socios"]:
            print(f"    Profs:   {row['Profesiones de los socios'][:80]}")
        if row["Objeto social"]:
            print(f"    Objeto:  {row['Objeto social'][:100]}")

    # Guardar Excel
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sociedades", index=False)
        ws = writer.sheets["Sociedades"]

        # Anchos personalizados (copiados de los ajustes manuales en el Excel).
        # Se aplican por NOMBRE de encabezado para no depender del orden de las columnas.
        anchos = {
            "Nombre de la sociedad": 52.5,
            "Tipo de sociedad": 15.1,
            "Tipo de acto": 11.2,
            "Descripción del acto": 209.5,
            "Nombres de los socios": 195.3,
            "Cargos de los socios": 255.7,
            "Domicilios particulares de los socios": 255.7,
            "Nombres de los socios jurídicos": 26.8,
            "CUIT de los socios jurídicos": 14.2,
            "Nombres de los apoderados": 28.5,
            "Escribano interviniente": 26.2,
            "Registro notarial": 21.8,
            "Domicilio electrónico de la sociedad": 30.5,
            "Domicilio electrónico de los socios": 74.2,
        }
        from openpyxl.utils import get_column_letter
        for idx, col_name in enumerate(df.columns, start=1):
            if col_name in anchos:
                ws.column_dimensions[get_column_letter(idx)].width = anchos[col_name]

        from openpyxl.styles import Alignment
        for row in ws.iter_rows(min_row=1, max_row=len(df) + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"

    # Checkpoint cumplió su función — borrar para no confundir corridas futuras
    if CHECKPOINT_FILE.exists():
        CHECKPOINT_FILE.unlink()
        logging.info("Checkpoint eliminado (corrida completa).")

    tiempo_total = time.time() - t0_total
    print(f"\n[OK] Excel guardado en: {OUTPUT_FILE}")
    print(f"[OK] Tiempo total: {_fmt_tiempo(tiempo_total)}")
    print(f"[OK] Costo estimado: ${_costo_estimado():.4f} USD")


if __name__ == "__main__":
    main()
